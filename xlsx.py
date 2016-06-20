# -*- coding: utf-8 -*-
"""

"""

import sys
reload(sys)
sys.setdefaultencoding('utf8')

from openpyxl import load_workbook
from openpyxl import Workbook
from xbuild import builDir
from openpyxl.cell import get_column_letter

class XlsxReader(object):
    def __init__(self, filename):
        self.workbook = load_workbook(filename='if00.xlsx', read_only=True)

    def readRow(self, sheet_index=0):
        # ws is now an IterableWorksheet
        worksheet = self.workbook[self.workbook.get_sheet_names()[0]]

        for row in worksheet.iter_rows():
            yield([cel.value for cel in row])


class XlsxWriter(object):
    def __init__(self, fileName, dirName='./'):
        self.filename = builDir(dirName, fileName)
        self.workbook = Workbook()
        self.sheet_name = 'sheet%s'
        self.sheet_count = 1
        self.workbook_is_open = 1
        self.sheet_is_open = 0

    def createSheet(self, title=''):
        if not title:
            title = self.sheet_name % self.sheet_count
        if self.sheet_count == 1:
            self.sheet = self.workbook.active
            self.sheet.title = title
        else:
            self.sheet = self.workbook.create_sheet(title=title)
        self.row_index = 0
        self.sheet_count += 1

    def writeRow(self, values):
        if not self.sheet_is_open:
            self.createSheet()
            self.sheet_is_open = 1
        col_num = len(values)
        for col, value in enumerate(values):
            _ = self.sheet.cell(column=col + 1,
                                row=self.row_index + 1,
                                value=value)
        self.row_index += 1

    def save(self):
        if self.workbook_is_open:
            self.workbook.save(filename = self.filename)
            self.workbook_is_open = 0

    def __exit__(self):
        """auto save"""
        if self.workbook_is_open:
            self.workbook.save(filename = self.filename)
            self.workbook_is_open = 0


if __name__ == '__main__':
    app = XlsxReader('if00.xlsx')
    app2 = XlsxWriter('a.xlsx')
    app2.createSheet()
    for values in app.readRow():
        app2.writeRow(values)
    app2.save()
